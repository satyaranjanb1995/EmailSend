#!/usr/bin/env python
# coding: utf-8

# In[1]:


import openpyxl


# In[2]:


import sys


# In[3]:


import smtplib


# In[4]:


# Open the spreadsheet and get the latest dues status.
wb = openpyxl.load_workbook('duesRecords.xlsx')
wb


# In[5]:


sheet = wb.sheetnames
sheet


# In[6]:


anotherSheet = wb.active


# In[7]:


max_col=anotherSheet.max_column
max_col


# In[8]:


lastCol = max_col
latestMonth = anotherSheet.cell(row=1, column=max_col).value


# In[9]:


latestMonth


# In[10]:


# Check each member's payment status.
unpaidMembers = {}
for i in range(2,anotherSheet.max_row+1):
    payment = anotherSheet.cell(row=i,column=anotherSheet.max_column).value
    if payment != 'paid':
        name = anotherSheet.cell(row=i,column=1).value
        email = anotherSheet.cell(row=i,column=2).value
        unpaidMembers[name]=email


# In[11]:


unpaidMembers


# In[12]:


email


# In[13]:


smtpObj = smtplib.SMTP('smtp-mail.outlook.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('satyaranjanbehera735@outlook.com', 'satya12345')

# Send out reminder emails.
for name, email in unpaidMembers.items():
    message="Please pay the due amount"
    subject = "June dues unpaid."
    body = "From: " + str(name) + "\n"
    body += "Email: " + str(email) + "\n"
    print('Sending email to %s...' % email)
    sendmailStatus = smtpObj.sendmail('satyaranjanbehera735@outlook.com', email, body)
    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email,sendmailStatus))
smtpObj.quit()


# In[14]:


sendmailStatus

